import base64
import io
from odoo import models, fields, api, exceptions

class ImportItemWizard(models.TransientModel):
    _name = 'material.tracker.import.wizard'
    _description = '导入料号向导'

    attachment_ids = fields.Many2many('ir.attachment', string='上传多文件', required=True)
    exist_action = fields.Selection([
        ('skip', '跳过'),
        ('overwrite', '覆盖')
    ], string='若项目已存在', default='skip', required=True)

    def action_import(self):
        self.ensure_one()
        try:
            import openpyxl
        except ImportError:
            raise exceptions.UserError('请安装 openpyxl 库：pip install openpyxl')

        if not self.attachment_ids:
            raise exceptions.UserError('请上传至少一个文件')

        total_items_created = 0
        total_files = len(self.attachment_ids)

        for attachment in self.attachment_ids:
            try:
                # 子项目名称来自文件名
                filename = attachment.name or '未知项目'
                sub_project_name = filename.rsplit('.', 1)[0]

                file_content = base64.b64decode(attachment.datas)
                workbook = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
                sheet = workbook.active

                items_to_create = []
                header_row_idx = -1
                col_map = {}
                parent_project_name = None  # 大项目名称从数据行"项目"列读取

                for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_strs = [str(cell).strip().replace('\n', '') if cell is not None else '' for cell in row]

                    # 识别表头行（含有"零件代号"等关键字）
                    if any('零件代号' in val or '料号' in val or '零件代码' in val for val in row_strs):
                        header_row_idx = idx
                        for i, val in enumerate(row_strs):
                            val_lower = val.lower()
                            if '零件代号L' in val: col_map['part_l'] = i
                            elif '数量L' in val: col_map['qty_l'] = i
                            elif '项目' in val: col_map['parent_proj'] = i
                            elif '下单人员' in val: col_map['order_person'] = i
                            elif '材料' in val: col_map['material'] = i
                            elif '备注' in val: col_map['remark'] = i
                            elif '工艺要求' in val or '表面处理' in val: col_map['process_req'] = i
                            elif '完成日期' in val or '日期' in val: col_map['due_date'] = i
                            elif '零件代号' in val or '料号' in val or '零件代码' in val:
                                if 'part_l' not in col_map:
                                    col_map['part_single'] = i
                            elif '数量' in val:
                                if 'qty_l' not in col_map:
                                    col_map['qty_single'] = i
                        break

                if header_row_idx == -1:
                    continue  # 跳过无效文件

                def get_val(row_data, key):
                    if key in col_map and col_map[key] < len(row_data):
                        return row_data[col_map[key]]
                    return None

                def get_qty(row_data, key):
                    val = get_val(row_data, key)
                    try:
                        return int(val) if val is not None and str(val).strip() else 1
                    except (ValueError, TypeError):
                        return 1

                def get_str(row_data, key):
                    val = get_val(row_data, key)
                    v_str = str(val).strip() if val is not None else ''
                    return v_str if v_str not in ('None', '') else False

                def get_date(row_data, key):
                    val = get_val(row_data, key)
                    if val is None:
                        return False
                    if hasattr(val, 'date'):
                        return val.date()
                    try:
                        from datetime import datetime
                        return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
                    except Exception:
                        return False

                # 读取大项目名称（从第一行数据读）
                first_data_row = next(sheet.iter_rows(min_row=header_row_idx + 2, values_only=True), None)
                if first_data_row and 'parent_proj' in col_map:
                    parent_project_name = get_str(first_data_row, 'parent_proj')

                # 创建或查找大项目（无 parent_id 的顶层项目）
                if parent_project_name:
                    parent_project = self.env['material.tracker.project'].search(
                        [('name', '=', parent_project_name), ('parent_id', '=', False)], limit=1
                    )
                    if not parent_project:
                        parent_project = self.env['material.tracker.project'].create({
                            'name': parent_project_name,
                        })
                else:
                    parent_project = False

                # 创建或查找子项目
                is_new_sub_project = False
                if parent_project:
                    sub_project = self.env['material.tracker.project'].search(
                        [('name', '=', sub_project_name), ('parent_id', '=', parent_project.id)], limit=1
                    )
                    if not sub_project:
                        sub_project = self.env['material.tracker.project'].create({
                            'name': sub_project_name,
                            'parent_id': parent_project.id,
                        })
                        is_new_sub_project = True
                else:
                    # 没有大项目时，直接作为顶层项目
                    sub_project = self.env['material.tracker.project'].search(
                        [('name', '=', sub_project_name), ('parent_id', '=', False)], limit=1
                    )
                    if not sub_project:
                        sub_project = self.env['material.tracker.project'].create({
                            'name': sub_project_name,
                        })
                        is_new_sub_project = True

                # --- 处理项目已存在的情况 ---
                if not is_new_sub_project:
                    if self.exist_action == 'skip':
                        continue  # 跳过该文件，处理下一个
                    elif self.exist_action == 'overwrite':
                        sub_project.item_ids.unlink()  # 删除该项目下的所有旧料号，准备重新导入

                for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
                    if not any(row):
                        continue

                    order_person = get_str(row, 'order_person')
                    material = get_str(row, 'material')
                    remark = get_str(row, 'remark')
                    process_req = get_str(row, 'process_req')
                    # 若没有独立的工艺要求列，则从"下单人员"列提取
                    if not process_req:
                        process_req = order_person
                    due_date = get_date(row, 'due_date')

                    item_vals = {
                        'project_id': sub_project.id,
                        'order_person': order_person,
                        'material': material,
                        'remark': remark,
                        'process_requirement': process_req,
                        'due_date': due_date,
                    }

                    if 'part_l' in col_map:
                        part_l = get_str(row, 'part_l')
                        if part_l:
                            item_vals.update({
                                'sku_code': part_l,
                                'qty': get_qty(row, 'qty_l'),
                            })
                            items_to_create.append(item_vals.copy())
                    elif 'part_single' in col_map:
                        part = get_str(row, 'part_single')
                        if part:
                            item_vals.update({
                                'sku_code': part,
                                'qty': get_qty(row, 'qty_single'),
                            })
                            items_to_create.append(item_vals.copy())

                if items_to_create:
                    self.env['material.tracker.item'].create(items_to_create)
                    total_items_created += len(items_to_create)

            except Exception as e:
                raise exceptions.UserError(f'处理文件 {attachment.name} 时失败：{str(e)}')

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
