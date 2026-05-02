import base64
import io
from odoo import models, fields, api, exceptions

class ImportOutsourcedWizard(models.TransientModel):
    _name = 'material.tracker.import.outsourced.wizard'
    _description = '导入外购件向导'

    attachment_ids = fields.Many2many('ir.attachment', string='上传文件', required=True)
    project_id = fields.Many2one('material.tracker.project', string='归属项目', required=True)

    def action_import(self):
        self.ensure_one()
        try:
            import openpyxl
        except ImportError:
            raise exceptions.UserError('请安装 openpyxl 库：pip install openpyxl')

        if not self.attachment_ids:
            raise exceptions.UserError('请上传文件')

        total_items_created = 0

        for attachment in self.attachment_ids:
            try:
                file_content = base64.b64decode(attachment.datas)
                workbook = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
                sheet = workbook.active

                items_to_create = []
                header_row_idx = -1
                col_map = {}

                for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_strs = [str(cell).strip() if cell is not None else '' for cell in row]

                    # 识别表头行
                    if any('名称' in val or '外购件' in val for val in row_strs):
                        header_row_idx = idx
                        for i, val in enumerate(row_strs):
                            if '名称' in val or '外购件' in val:
                                col_map['name'] = i
                            elif '数量' in val:
                                col_map['qty'] = i
                        break

                if header_row_idx == -1 or 'name' not in col_map:
                    continue  # 跳过无效文件

                for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
                    if not any(row):
                        continue
                    
                    name = str(row[col_map['name']]).strip() if col_map['name'] < len(row) and row[col_map['name']] else None
                    if not name:
                        continue
                    
                    qty_val = row[col_map['qty']] if 'qty' in col_map and col_map['qty'] < len(row) else 1
                    try:
                        qty = int(qty_val) if qty_val is not None else 1
                    except (ValueError, TypeError):
                        qty = 1

                    items_to_create.append({
                        'project_id': self.project_id.id,
                        'name': name,
                        'qty': qty,
                    })

                if items_to_create:
                    self.env['material.tracker.outsourced.item'].create(items_to_create)
                    total_items_created += len(items_to_create)

            except Exception as e:
                raise exceptions.UserError(f'处理文件 {attachment.name} 时失败：{str(e)}')

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
